#!/usr/bin/env python3
"""Reset the publications pipeline for a new author.

Wipes all personal data (papers, citations, generated bib/tex) and
replaces main.tex with the blank template so you can start fresh.

Usage:
    python init_new_author.py [--overleaf-url URL] [--yes]

Options:
    --overleaf-url URL   Git URL of your new Overleaf project.
                         If omitted, instructions are printed instead.
    --yes                Skip the confirmation prompt.

After running this script:
    1. Edit config.py to set your name and Google Scholar user ID.
    2. Run: python rebuild_tex.py   (generates Wzmn.bib and main.tex from scratch)
    3. Run: python update.py        (full Scholar fetch + push to Overleaf)
"""

import argparse
import json
import os
import shutil
import subprocess
import sys

FILE_DIR = os.path.dirname(os.path.abspath(__file__))

# The personal data this script removes, in the order main() prints it. Single
# source of truth: main()'s summary and the tests both read this, so a file
# added to the pipeline cannot be silently left behind by one of them.
PERSONAL_FILES = (
    ("citations.csv",          "per-paper citation counts"),
    ("papers.csv",             "the publications table"),
    ("orig.bib",               "raw BibTeX entries"),
    ("profile_stats.json",     "Scholar totals (citations, h-index)"),
    ("overleaf/Wzmn.bib",      "generated bibliography"),
    ("identity.json",          "harvested paper identifiers"),
    ("resolve_attempts.json",  "resolve attempt counters"),
    (".pipeline_state.json",   "step completion state"),
    ("WORKLIST.md",            "open items report"),
    ("tmp.csv",                "temporary Scholar fetch file"),
)

# Deleted outright rather than emptied: machine-owned state, rebuilt on demand.
_DELETED_OUTRIGHT = ("resolve_attempts.json", "identity.json",
                     ".pipeline_state.json", "WORKLIST.md", "tmp.csv")

XLSX_NAME = "Contributions_table.xlsx"

# Preferred first. templates/main.tex is tracked in this repository; the
# submodule copy is only reachable while overleaf/ still points at the original
# author's project, which stops being true the moment a fork repoints it.
_TEMPLATE_CANDIDATES = (os.path.join("templates", "main.tex"),
                        os.path.join("overleaf", "template.tex"))


def _confirm(prompt: str) -> bool:
    return input(prompt + " [y/N] ").strip().lower() == "y"


def wipe_citations_csv(root=None):
    from citations_io import write_citation_rows
    root = root or FILE_DIR
    write_citation_rows([], os.path.join(root, "citations.csv"))
    print("  Cleared citations.csv")


def wipe_profile_stats(root=None):
    root = root or FILE_DIR
    with open(os.path.join(root, "profile_stats.json"), "w") as f:
        json.dump({"citations": 0, "h_index": 0}, f)
    print("  Reset profile_stats.json")


def wipe_orig_bib(root=None):
    root = root or FILE_DIR
    open(os.path.join(root, "orig.bib"), "w").close()
    print("  Cleared orig.bib")


def wipe_wzmn_bib(root=None):
    root = root or FILE_DIR
    path = os.path.join(root, "overleaf", "Wzmn.bib")
    if os.path.exists(path):
        open(path, "w").close()
        print("  Cleared overleaf/Wzmn.bib")


def wipe_tmp_csv(root=None):
    root = root or FILE_DIR
    path = os.path.join(root, "tmp.csv")
    if os.path.exists(path):
        os.remove(path)
        print("  Deleted tmp.csv")


def wipe_resolve_attempts(root=None):
    root = root or FILE_DIR
    for name in _DELETED_OUTRIGHT:
        if name == "tmp.csv":
            continue        # handled by wipe_tmp_csv, which reports it
        path = os.path.join(root, name)
        if os.path.exists(path):
            os.remove(path)
            print(f"  Deleted {name}")


def wipe_contributions_xlsx(root=None):
    """Empty the publications table, keeping its columns.

    Handles both formats: papers.csv is the current source of truth, and the
    xlsx is cleared too so a fork that has not migrated starts clean either way.
    """
    root = root or FILE_DIR
    papers_csv = os.path.join(root, "papers.csv")
    if os.path.exists(papers_csv):
        import pandas as pd
        from table_io import write_table
        df = pd.read_csv(papers_csv, dtype=str, nrows=0)
        write_table(df, papers_csv)
        print("  Cleared papers.csv (columns preserved)")

    xlsx_path = os.path.join(root, XLSX_NAME)
    if os.path.exists(xlsx_path):
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        wb.save(xlsx_path)
        print(f"  Cleared {XLSX_NAME} (header row preserved)")


def find_template(root=None):
    """Path to the blank CV template, or None if neither copy is present."""
    root = root or FILE_DIR
    for rel in _TEMPLATE_CANDIDATES:
        path = os.path.join(root, rel)
        if os.path.exists(path):
            return path
    return None


def reset_main_tex(root=None):
    """Replace overleaf/main.tex with the blank template. True on success."""
    root = root or FILE_DIR
    template = find_template(root)
    if template is None:
        print("  ERROR: no CV template found. Expected one of: "
              + ", ".join(_TEMPLATE_CANDIDATES))
        return False
    main_tex = os.path.join(root, "overleaf", "main.tex")
    os.makedirs(os.path.dirname(main_tex), exist_ok=True)
    shutil.copy2(template, main_tex)
    print(f"  Reset overleaf/main.tex from {os.path.relpath(template, root)}")
    return True


def replace_overleaf_submodule(overleaf_url: str, root=None):
    """Point the overleaf/ submodule at a new Overleaf project. True on success."""
    root = root or FILE_DIR
    print(f"\nReplacing overleaf/ submodule → {overleaf_url}")

    # Cleanup of whatever state the clone is in; each of these legitimately
    # fails in some states, so none is checked. Only `submodule add` has to
    # succeed.
    #
    # `rm --cached` is not redundant with `rm -f`. When the submodule is
    # committed but not initialised -- the normal state of a fork, since cloning
    # the original needs the previous author's credentials -- `rm -f` cannot
    # remove it, the index entry survives, and `submodule add` then fails with
    # "'overleaf' already exists in the index".
    for cmd in (["submodule", "deinit", "-f", "overleaf"],
                ["rm", "-f", "overleaf"],
                ["rm", "-rf", "--cached", "overleaf"]):
        subprocess.run(["git", "-C", root] + cmd, capture_output=True, text=True)

    # After deinit and rm, so neither is looking for a gitdir that has just
    # been deleted underneath it.
    modules_dir = os.path.join(root, ".git", "modules", "overleaf")
    if os.path.exists(modules_dir):
        shutil.rmtree(modules_dir)
    leftover = os.path.join(root, "overleaf")
    if os.path.isdir(leftover):
        # `submodule add` refuses a path that already exists.
        shutil.rmtree(leftover, ignore_errors=True)

    result = subprocess.run(["git", "-C", root, "submodule", "add",
                             overleaf_url, "overleaf"],
                            capture_output=True, text=True)
    if result.returncode != 0:
        print(f"  ERROR running submodule add:\n{result.stderr.strip()}")
        return False
    print("  overleaf/ now points to your Overleaf project")
    print("  Tip: run `git commit -m 'chore: switch to new Overleaf project'` "
          "to record this change")
    return True


def print_overleaf_instructions():
    print("""
  To connect your own Overleaf project:
    1. Create a new project on overleaf.com.
    2. In Overleaf: Menu → Git → copy the project URL, and generate a token at
       Account Settings → Git integration. The URL you clone with is
       https://git:<token>@git.overleaf.com/<project-id>
    3. Run:
         python init_new_author.py --overleaf-url <that-url>
    4. Then run `python rebuild_tex.py` to populate overleaf/ with your files.
""")


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Reset the publications pipeline for a new author",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--overleaf-url", metavar="URL",
                        help="Git URL of your new Overleaf project")
    parser.add_argument("--yes", action="store_true",
                        help="Skip confirmation prompt")
    args = parser.parse_args(argv)

    print("This will erase all personal data from the pipeline:")
    for name, what in PERSONAL_FILES:
        print(f"  • {name:<24} ({what})")
    print(f"  • {'overleaf/main.tex':<24} → replaced with the blank template")
    if args.overleaf_url:
        print(f"  • {'overleaf/ submodule':<24} → {args.overleaf_url}")
    print()

    if not args.yes and not _confirm("Proceed?"):
        print("Aborted.")
        return 0

    wipe_citations_csv()
    wipe_profile_stats()
    wipe_orig_bib()
    wipe_wzmn_bib()
    wipe_tmp_csv()
    wipe_resolve_attempts()
    wipe_contributions_xlsx()

    # The submodule swap first: it deletes overleaf/ and re-clones the new
    # project over it, so a main.tex written before it does not survive. A new
    # Overleaf project is empty, which left the fork with no CV at all -- and a
    # "Done" telling it to run rebuild_tex.py against nothing.
    ok = True
    if args.overleaf_url:
        ok = replace_overleaf_submodule(args.overleaf_url)
    else:
        print_overleaf_instructions()
    ok = reset_main_tex() and ok

    if not ok:
        # Exits non-zero: the personal data is gone either way, but the setup is
        # incomplete, and printing "Done" over a failed submodule swap sent a
        # fork on to the next step with overleaf/ still pointing at this author.
        print("\nSomething above did not complete. Fix it before continuing.")
        return 1

    print("\nDone. Next steps:")
    print("  1. Edit config.py  — set AUTHOR_NAME and SCHOLAR_USER_ID")
    print("  2. python rebuild_tex.py   — generate initial Wzmn.bib and main.tex")
    print("  3. python update.py        — fetch from Scholar and push to Overleaf")
    return 0


if __name__ == "__main__":
    sys.exit(main())
