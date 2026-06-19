#!/usr/bin/env python3
"""One-command update for the publications pipeline.

Steps (each auto-skips if output is already newer than its inputs):
  1. Refresh citations.csv from Google Scholar
  2. Add new papers (in Scholar, not in xlsx) to Contributions_table.xlsx
  3. Resolve arXiv entries in orig.bib to published BibTeX (in-place);
     also resolve xlsx entries with no Bib key and add them to orig.bib.
  4. Build wzmn.bib from orig.bib + xlsx metadata (via build_bib.py)
  5. Rebuild example.tex with updated \\nocite{} blocks (via rebuild_tex.py)
  6. Commit changed files and push to origin (GitHub) and overleaf remotes

Usage:
    python update.py [--dry-run] [--force] [--no-push]
                     [--skip-fetch] [--skip-xlsx] [--skip-resolve] [--skip-publications]
                     [--skip-tex] [--fetch-age HOURS] [--user SCHOLAR_USER_ID]
"""

import argparse
import csv
import difflib
import os
import re
import subprocess
import sys
import time

import openpyxl

FILE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, FILE_DIR)

from resolve_arxiv import (
    _PUBLISHED_SOURCES,
    _get_arxiv_id,
    gen_key,
    get_arxiv_entries,
    resolve,
    update_bib_inplace,
)

CITATIONS_CSV = os.path.join(FILE_DIR, "citations.csv")
XLSX_PATH     = os.path.join(FILE_DIR, "Contributions_table.xlsx")
BIB_PATH      = os.path.join(FILE_DIR, "orig.bib")
OVERLEAF_DIR  = os.path.join(FILE_DIR, "overleaf")
WZMN_BIB      = os.path.join(OVERLEAF_DIR, "Wzmn.bib")
TEX_PATH      = os.path.join(OVERLEAF_DIR, "main.tex")

# xlsx column indices (0-based matching ws.iter_rows)
COL_ID      = 0   # Time of publish ID
COL_VENUE   = 2
COL_NAME    = 3
COL_BIB     = 4
COL_AUTHORS = 5
COL_YEAR    = 6
COL_PAPER   = 26


# ── Helpers ────────────────────────────────────────────────────────────────────

def _simplify(text: str) -> str:
    return re.sub(r'[\W_]+', '', text.lower().strip())


def _mtime(path: str) -> float:
    """Return file mtime as float, or 0.0 if the file doesn't exist."""
    try:
        return os.path.getmtime(path)
    except OSError:
        return 0.0


def _age_hours(path: str) -> float:
    """Return file age in hours, or inf if the file doesn't exist."""
    mt = _mtime(path)
    return (time.time() - mt) / 3600 if mt else float("inf")


def _parse_scholar_csv(csv_path: str) -> list:
    """Parse citations.csv (3-row-per-paper) into list of dicts."""
    papers = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    i = 1  # skip header row
    while i < len(rows):
        r0 = rows[i]     if i     < len(rows) else []
        r1 = rows[i + 1] if i + 1 < len(rows) else []
        r2 = rows[i + 2] if i + 2 < len(rows) else []
        title   = r0[0].strip() if r0 else ""
        year    = r0[2].strip() if len(r0) > 2 else ""
        authors = r1[0].strip() if r1 else ""
        venue   = r2[0].strip() if r2 else ""
        if title:
            papers.append({"title": title, "year": year, "authors": authors, "venue": venue})
        i += 3
    return papers


def _get_max_id(ws) -> int:
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[COL_ID]
        if isinstance(v, (int, float)) and v > 0:
            max_id = max(max_id, int(v))
    return max_id


def _last_data_row(ws) -> int:
    """Return the row number of the last row with a non-empty Name cell."""
    last = 1
    for row in ws.iter_rows(min_row=2):
        if row[COL_NAME].value:
            last = row[0].row
    return last



# ── Step 1 ─────────────────────────────────────────────────────────────────────

def step1_fetch(dry_run: bool, user: str | None = None) -> None:
    print("[Step 1] Fetching Scholar profile → citations.csv")
    if dry_run:
        print("  (dry-run: skipped)")
        return
    cmd = [sys.executable, os.path.join(FILE_DIR, "fetch_citations.py")]
    if user:
        cmd.append(user)
    subprocess.run(cmd, check=True)
    print("  Done.")


# ── Step 2 ─────────────────────────────────────────────────────────────────────

def step2_add_new_papers(dry_run: bool) -> int:
    """Return number of new papers added (or that would be added)."""
    print("\n[Step 2] Checking for new papers not in Contributions_table.xlsx")
    wb     = openpyxl.load_workbook(XLSX_PATH)
    ws     = wb.active
    papers = _parse_scholar_csv(CITATIONS_CSV)

    # Build simplified→original mapping for both matching and display
    simp_to_orig: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[COL_NAME]
        if name:
            simp_to_orig[_simplify(str(name))] = str(name)
    known = list(simp_to_orig)

    new_papers = [
        p for p in papers
        if not difflib.get_close_matches(_simplify(p["title"]), known, n=1, cutoff=0.85)
    ]
    if not new_papers:
        print("  No new papers found.")
        return 0

    print(f"  {len(new_papers)} new paper(s):")
    max_id = _get_max_id(ws)
    for i, p in enumerate(new_papers, 1):
        year_val = int(p["year"]) if p["year"].isdigit() else p["year"]
        simp = _simplify(p["title"])
        best_k = max(known, key=lambda k: difflib.SequenceMatcher(None, simp, k).ratio()) if known else ""
        best_score = difflib.SequenceMatcher(None, simp, best_k).ratio() if best_k else 0
        hint = (f"\n        ↑ closest in xlsx: {simp_to_orig[best_k][:60]!r} ({best_score:.0%})"
                if best_score > 0.60 else "")
        print(f"    [{year_val}] {p['title'][:70]}{hint}")

    if not dry_run:
        insert_at = _last_data_row(ws) + 1
        ws.insert_rows(insert_at, len(new_papers))
        for i, p in enumerate(new_papers):
            year_val = int(p["year"]) if p["year"].isdigit() else p["year"]
            row_data = [None] * 37
            row_data[COL_ID]      = max_id + i + 1
            row_data[COL_VENUE]   = p["venue"]
            row_data[COL_NAME]    = p["title"]
            row_data[COL_AUTHORS] = p["authors"]
            row_data[COL_YEAR]    = year_val
            row_data[COL_PAPER]   = 1
            for j, val in enumerate(row_data):
                ws.cell(row=insert_at + i, column=j + 1).value = val
        wb.save(XLSX_PATH)
        print(f"  Saved {len(new_papers)} new row(s) to Contributions_table.xlsx")
    else:
        print("  (dry-run: xlsx not modified)")
    return len(new_papers)


# ── Step 3 ─────────────────────────────────────────────────────────────────────

def _get_xlsx_missing(bib_text: str) -> list:
    """Return xlsx rows with no Bib key (or key absent from bib), with gen_key applied."""
    from bib_utils import parse_bibtex as _parse_bib
    try:
        from bib_utils import read_df
        df = read_df()
    except Exception as exc:
        print(f"  Warning: could not read xlsx: {exc}", file=sys.stderr)
        return []
    existing_keys = {e["item_name"] for e in _parse_bib(bib_text)}
    missing = []
    for _, row in df.iterrows():
        bib_key = str(row.get("Bib", "")).strip()
        name    = str(row.get("Name", "")).strip()
        authors = str(row.get("Authors", "") or "").strip()
        if authors.lower() == "nan":
            authors = ""
        year    = str(int(row.get("year", 0) or 0))
        if not name:
            continue
        key_missing = not bib_key or bib_key.lower() in ("nan", "none") or bib_key not in existing_keys
        if key_missing:
            if bib_key and bib_key.lower() not in ("nan", "none"):
                new_key = bib_key  # key was set in xlsx but not yet in orig.bib; preserve it
            elif authors:
                new_key = gen_key(authors, year, name)
            else:
                new_key = f"unknown{year}{_simplify(name)[:10]}"
            missing.append({"title": name, "authors": authors, "year": year,
                            "item_name": new_key, "content": ""})
    return missing


def step3_resolve(dry_run: bool) -> tuple:
    print("\n[Step 3] Resolving BibTeX entries and updating orig.bib")
    with open(BIB_PATH) as f:
        bib_text = f.read()

    # Part A: existing arXiv entries in orig.bib
    arxiv_entries = get_arxiv_entries(bib_text)
    print(f"  {len(arxiv_entries)} arXiv entries to check in orig.bib...")
    updates = []
    for entry in arxiv_entries:
        key      = entry["item_name"]
        arxiv_id = _get_arxiv_id(entry)
        label    = f"arXiv:{arxiv_id}" if arxiv_id else "(no arXiv ID)"
        print(f"    [{key[:40]}] {label:<22}", end=" ", flush=True)
        bib, source = resolve(entry["title"], arxiv_id, key, entry.get("content", ""))
        print(f"→ {source}")
        updates.append((key, bib, source))
        time.sleep(0.5)

    # Part B: xlsx entries with no Bib key
    missing_entries = _get_xlsx_missing(bib_text)
    new_entries = []
    not_found = []
    if missing_entries:
        print(f"\n  {len(missing_entries)} xlsx entry(ies) with no BibTeX key...")
    for entry in missing_entries:
        key = entry["item_name"]
        print(f"    [{key:<40}]", end=" ", flush=True)
        bib, source = resolve(entry["title"], None, key, "")
        print(f"→ {source}")
        if bib:
            new_entries.append((key, bib))
        else:
            not_found.append((entry["title"][:70], key))
        time.sleep(0.5)

    # Also track arXiv entries that couldn't be resolved to any bib at all
    for key, bib, source in updates:
        if not bib:
            title = next((e["title"] for e in arxiv_entries if e["item_name"] == key), key)
            not_found.append((title[:70], key))

    # Summary before writing
    upgraded = sum(1 for _, _, src in updates if src in _PUBLISHED_SOURCES)
    print(f"\n  {upgraded}/{len(arxiv_entries)} arXiv entries have a published version")
    print(f"  {len(new_entries)} new entries to append to orig.bib")

    n_still_arxiv = len(arxiv_entries) - upgraded
    if dry_run:
        print("  (dry-run: orig.bib not modified)")
        return upgraded, len(new_entries), n_still_arxiv, not_found

    new_bib_text, n_replaced, n_appended = update_bib_inplace(bib_text, updates, new_entries)

    # Write xlsx before orig.bib so orig.bib ends up as the newest file,
    # which lets the mtime-based auto-skip correctly detect "step 3 already ran".
    if new_entries:
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active
        resolved_keys = {key for key, _ in new_entries}
        for row in ws.iter_rows(min_row=2):
            name_cell    = row[COL_NAME]
            bib_cell     = row[COL_BIB]
            authors_cell = row[COL_AUTHORS]
            year_cell    = row[COL_YEAR]
            if not name_cell.value or bib_cell.value:
                continue
            name    = str(name_cell.value).strip()
            authors = str(authors_cell.value or "").strip()
            year    = str(int(year_cell.value) if isinstance(year_cell.value, (int, float)) else year_cell.value or 0)
            candidate_key = (gen_key(authors, year, name) if authors
                             else f"unknown{year}{_simplify(name)[:10]}")
            if candidate_key in resolved_keys:
                bib_cell.value = candidate_key
        wb.save(XLSX_PATH)
        print(f"  Updated Bib keys in Contributions_table.xlsx")

    # Always write orig.bib (even if content unchanged) to update its mtime,
    # marking step 3 as "completed against current inputs" for the next auto-skip check.
    with open(BIB_PATH, "w") as f:
        f.write(new_bib_text)
    if n_replaced or n_appended:
        print(f"  Replaced {n_replaced} entries, appended {n_appended} entries → orig.bib updated")
    else:
        print(f"  No bib changes — orig.bib mtime touched")
    return upgraded, n_appended, n_still_arxiv, not_found


# ── Step 4 ─────────────────────────────────────────────────────────────────────

def step4_build_bib(dry_run: bool):
    print("\n[Step 4] Building wzmn.bib")
    if dry_run:
        print("  (dry-run: skipped)")
        return None
    import build_bib
    return build_bib.main()


# ── Step 5 ─────────────────────────────────────────────────────────────────────

def step5_rebuild_tex(dry_run: bool, cats) -> None:
    print("\n[Step 5] Rebuilding example.tex")
    if dry_run:
        print("  (dry-run: skipped)")
        return
    import rebuild_tex
    rebuild_tex.main(cats)


# ── Step 6 ─────────────────────────────────────────────────────────────────────

_OUTER_FILES = [
    "citations.csv",
    "profile_stats.json",
    "Contributions_table.xlsx",
    "orig.bib",
    "overleaf",  # submodule pointer
]
_OVERLEAF_FILES = ["main.tex", "Wzmn.bib"]


def _git_commit_and_push(repo_dir: str, files: list[str], message: str, remote: str) -> bool:
    """Stage files, commit if changed, push. Returns True on success."""
    existing = [f for f in files if os.path.exists(os.path.join(repo_dir, f))]
    subprocess.run(["git", "-C", repo_dir, "add", "--"] + existing, capture_output=True)

    diff = subprocess.run(["git", "-C", repo_dir, "diff", "--cached", "--quiet"], capture_output=True)
    if diff.returncode == 0:
        print(f"  [{remote}] Nothing to commit.")
    else:
        result = subprocess.run(
            ["git", "-C", repo_dir, "commit", "-m", message],
            capture_output=True, text=True,
        )
        if result.returncode != 0:
            print(f"  [{remote}] Commit failed: {result.stderr.strip()}")
            return False
        print(f"  [{remote}] Committed.")

    print(f"  [{remote}] Pushing…", end=" ", flush=True)
    push = subprocess.run(["git", "-C", repo_dir, "push", remote], capture_output=True, text=True)
    if push.returncode == 0:
        print("ok")
        return True
    print(f"FAILED\n    {push.stderr.strip()}")
    return False


def step6_push(dry_run: bool) -> None:
    print("\n[Step 6] Committing and pushing to Overleaf + GitHub")
    if dry_run:
        print("  (dry-run: skipped)")
        return

    # Push submodule (overleaf/) → Overleaf
    _git_commit_and_push(
        OVERLEAF_DIR, _OVERLEAF_FILES,
        "chore: auto-update publications pipeline output",
        "origin",
    )

    # Push outer repo (publications/) → GitHub
    _git_commit_and_push(
        FILE_DIR, _OUTER_FILES,
        "chore: auto-update publications pipeline output",
        "origin",
    )


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Update the publications pipeline end-to-end",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Steps auto-skip when their output is already newer than their inputs.\nUse --force to bypass all auto-skip checks.",
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Show what would change without modifying files")
    parser.add_argument("--force", action="store_true",
                        help="Run all steps regardless of file mtimes")
    parser.add_argument("--fetch-age", type=float, default=24.0, metavar="HOURS",
                        help="Re-fetch citations.csv if older than this many hours (default: 24)")
    parser.add_argument("--skip-fetch",        action="store_true", help="Skip step 1")
    parser.add_argument("--skip-xlsx",         action="store_true", help="Skip step 2")
    parser.add_argument("--skip-resolve",      action="store_true", help="Skip step 3")
    parser.add_argument("--skip-publications", action="store_true", help="Skip step 4")
    parser.add_argument("--skip-tex",          action="store_true", help="Skip step 5")
    parser.add_argument("--no-push",           action="store_true", help="Skip step 6 (commit + push)")
    parser.add_argument("--user", default=None,
                        help="Google Scholar user ID (passed to fetch_citations.py)")
    args = parser.parse_args()

    n_added = 0
    n_upgraded = n_appended = n_still_arxiv = 0
    not_found: list = []
    cats = None

    # Step 1 — re-fetch if citations.csv is stale
    csv_age = _age_hours(CITATIONS_CSV)
    if args.skip_fetch:
        print("[Step 1] Skipped.")
    elif not args.force and csv_age < args.fetch_age:
        print(f"[Step 1] Auto-skipped — citations.csv is {csv_age:.1f}h old "
              f"(threshold: {args.fetch_age}h, use --force to override).")
    else:
        step1_fetch(args.dry_run, args.user)

    # Step 2 — add new papers (fast and idempotent; always run unless explicitly skipped)
    if args.skip_xlsx:
        print("\n[Step 2] Skipped.")
    else:
        n_added = step2_add_new_papers(args.dry_run)

    # Step 3 — resolve bib entries; auto-skip if orig.bib is newer than both inputs
    bib_stale = _mtime(BIB_PATH) < max(_mtime(XLSX_PATH), _mtime(CITATIONS_CSV))
    if args.skip_resolve:
        print("\n[Step 3] Skipped.")
    elif not args.force and not bib_stale:
        print("\n[Step 3] Auto-skipped — orig.bib is newer than xlsx and citations.csv.")
    else:
        n_upgraded, n_appended, n_still_arxiv, not_found = step3_resolve(args.dry_run)

    # Step 4 — build wzmn.bib; auto-skip if it is newer than both inputs
    wzmn_stale = _mtime(WZMN_BIB) < max(_mtime(BIB_PATH), _mtime(XLSX_PATH))
    if args.skip_publications:
        print("\n[Step 4] Skipped.")
    elif not args.force and not wzmn_stale:
        print("\n[Step 4] Auto-skipped — wzmn.bib is newer than orig.bib and xlsx.")
    else:
        cats = step4_build_bib(args.dry_run)

    # Step 5 — rebuild example.tex; auto-skip if it is newer than wzmn.bib
    tex_stale = _mtime(TEX_PATH) < _mtime(WZMN_BIB)
    if args.skip_tex:
        print("\n[Step 5] Skipped.")
    elif not args.force and not tex_stale:
        print("\n[Step 5] Auto-skipped — example.tex is newer than wzmn.bib.")
    else:
        step5_rebuild_tex(args.dry_run, cats)

    # Step 6 — commit + push to origin and overleaf
    if args.no_push:
        print("\n[Step 6] Skipped (--no-push).")
    else:
        step6_push(args.dry_run)

    print("\n" + "═" * 52)
    print(f"  Step 2: {n_added} new paper(s) added to xlsx")
    print(f"  Step 3: {n_upgraded} arXiv → published  |  "
          f"{n_appended} new entries appended  |  "
          f"{n_still_arxiv} still arXiv")
    if not_found:
        print(f"\n  WARNING: {len(not_found)} paper(s) need manual bib lookup:")
        for title, key in not_found:
            print(f"    [{key}] {title}")
    print("═" * 52)


if __name__ == "__main__":
    main()
